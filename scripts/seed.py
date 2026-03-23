#!/usr/bin/env python3
"""
seed.py — Seed completo do PortalTradeHub.

FASE 1 (direct DB via pymysql):
  • Limpar tabelas (--reset)
  • Dados-mestre: banks, products, departments, teams,
    error_impacts, error_detected_by, error_origins,
    tutoria_error_categories, activities, error_types
  • Utilizadores de teste (password: test123)

FASE 2 (API via requests):
  • Cursos + Lições
  • Planos de Formação + Inscrições
  • Desafios + Submissões
  • Erros de Tutoria + Planos de Ação + Fichas de Aprendizagem
  • Cápsulas Formativas
  • Sensos + Erros Internos
  • Chamados de Suporte
  • Feedback Surveys + Respostas + Ações
  • FAQs do Chat
  • Team Services

Uso:
  python scripts/seed.py                      # seed completo (mantém dados existentes)
  python scripts/seed.py --reset              # limpa tudo e volta a semear
  python scripts/seed.py --reset --url http://localhost:8100
  python scripts/seed.py --skip-master        # só fase 2 (dados-mestre já existem)
  python scripts/seed.py --skip-operational   # só fase 1 (sem chamadas API)
  python scripts/seed.py --dry-run            # mostra contagens, sem inserir

Requisitos: pip install pymysql openpyxl requests
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "docs" / "Exemplo dados reais"
XLSX_FILE = DATA_DIR / "EXEMPLOS.xlsx"

# ── Args ───────────────────────────────────────────────────────────────────────
ap = argparse.ArgumentParser(description="Seed completo do PortalTradeHub")
ap.add_argument("--reset",            action="store_true", help="Truncar todas as tabelas antes de semear")
ap.add_argument("--dry-run",          action="store_true", help="Só mostra contagens, sem inserir")
ap.add_argument("--skip-master",      action="store_true", help="Saltar fase 1 (dados-mestre + utilizadores)")
ap.add_argument("--skip-operational", action="store_true", help="Saltar fase 2 (dados operacionais via API)")
ap.add_argument("--skip-excel",       action="store_true", help="Saltar seeds do Excel (activities, error_types)")
ap.add_argument("--host",  default="localhost",        help="Host MySQL (default: localhost)")
ap.add_argument("--port",  type=int, default=3307,     help="Porta MySQL (default: 3307, Docker expõe aqui)")
ap.add_argument("--url",   default="http://localhost:8100", help="URL do backend (default: http://localhost:8100)")
ap.add_argument("--verbose", action="store_true",      help="Mostrar detalhes de cada chamada API")
args = ap.parse_args()

DRY = args.dry_run
V   = args.verbose

# ══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════════════

def _load_env_file() -> dict[str, str]:
    """Lê ambos backend/.env e .env raiz; a raiz tem prioridade para vars MySQL."""
    result: dict[str, str] = {}
    for env_file in [ROOT / "backend" / ".env", ROOT / ".env"]:
        if env_file.exists():
            for line in env_file.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    k = k.strip()
                    if k not in result:  # primeiro a encontrar a chave ganha
                        result[k] = v.strip().strip('"').strip("'")
    return result


ENV = _load_env_file()
DB_PASS = os.getenv("MYSQL_ROOT_PASSWORD", ENV.get("MYSQL_ROOT_PASSWORD", "dev_root_pass"))
DB_USER = os.getenv("MYSQL_USER", ENV.get("MYSQL_USER", "tradehub_user"))
DB_USER_PASS = os.getenv("MYSQL_PASSWORD", ENV.get("MYSQL_PASSWORD", "dev_pass"))
DB_NAME = os.getenv("MYSQL_DATABASE", ENV.get("MYSQL_DATABASE", "tradehub_db"))

print(f"\n{'='*65}")
print(f"  PortalTradeHub — Seed Completo")
print(f"  DB  : {args.host}:{args.port}/{DB_NAME}")
print(f"  API : {args.url}")
if DRY:
    print("  MODE: DRY-RUN (sem writes)")
if args.reset:
    print("  RESET: todas as tabelas serão limpas")
print(f"{'='*65}\n")


# ══════════════════════════════════════════════════════════════════════════════
# FASE 1 — DADOS-MESTRE + UTILIZADORES (direct DB)
# ══════════════════════════════════════════════════════════════════════════════

if not args.skip_master:
    try:
        import pymysql
    except ImportError:
        sys.exit("ERROR: pip install pymysql")

    try:
        conn = pymysql.connect(
            host=args.host, port=args.port,
            user="root", password=DB_PASS,
            database=DB_NAME, charset="utf8mb4",
            autocommit=False,
        )
    except Exception:
        # Fallback: try with app user
        try:
            conn = pymysql.connect(
                host=args.host, port=args.port,
                user=DB_USER, password=DB_USER_PASS,
                database=DB_NAME, charset="utf8mb4",
                autocommit=False,
            )
        except Exception as exc:
            sys.exit(f"ERROR: não foi possível ligar à DB: {exc}")

    cur = conn.cursor()

    # ── Helpers DB ─────────────────────────────────────────────────────────

    def _existing_set(table: str, col: str = "name") -> set[str]:
        cur.execute(f"SELECT `{col}` FROM `{table}`")
        return {r[0] for r in cur.fetchall()}

    def _upsert(table: str, cols: list, rows: list, label: str = "") -> int:
        if not rows:
            print(f"  [--] {label or table}: nada a inserir")
            return 0
        if DRY:
            print(f"  [dry] {label or table}: {len(rows)} linhas")
            return len(rows)
        ph = ", ".join(["%s"] * len(cols))
        col_str = ", ".join(f"`{c}`" for c in cols)
        sql = f"INSERT IGNORE INTO `{table}` ({col_str}) VALUES ({ph})"
        cur.executemany(sql, rows)
        conn.commit()
        inserted = cur.rowcount
        skipped = len(rows) - inserted
        print(f"  [OK] {label or table}: {inserted} inseridos, {skipped} já existiam")
        return inserted

    def _insert_names(table: str, names: list, extra: dict | None = None, label: str = "") -> int:
        exist = _existing_set(table)
        new = [n for n in names if n not in exist]
        if not new:
            print(f"  [--] {label or table}: todos os {len(names)} já existem")
            return 0
        extra = extra or {}
        cols = ["name", "is_active"] + list(extra.keys())
        rows = [(n, 1) + tuple(extra.values()) for n in new]
        return _upsert(table, cols, rows, label or table)

    # ── RESET ──────────────────────────────────────────────────────────────

    if args.reset:
        print("── Reset: limpando todas as tabelas ─────────────────────────────")
        if DRY:
            print("  [dry] TRUNCATE todas as tabelas")
        else:
            # Listar todas as tabelas da DB (excluir migrações e vistas DW)
            cur.execute(
                "SELECT table_name FROM information_schema.tables "
                "WHERE table_schema = %s AND table_type = 'BASE TABLE' "
                "AND table_name NOT IN ('_migrations', 'alembic_version')",
                (DB_NAME,)
            )
            all_tables = [r[0] for r in cur.fetchall()
                          if not r[0].startswith("dw_")]
            cur.execute("SET FOREIGN_KEY_CHECKS = 0")
            truncated = 0
            for tbl in all_tables:
                try:
                    cur.execute(f"TRUNCATE TABLE `{tbl}`")
                    truncated += 1
                    if V:
                        print(f"  [TRUNCATE] {tbl}")
                except Exception:
                    pass
            cur.execute("SET FOREIGN_KEY_CHECKS = 1")
            conn.commit()
            print(f"  [OK] {truncated} tabelas limpas.\n")

    # ── SECÇÃO 1: Dados fixos (dos PNGs) ───────────────────────────────────

    print("── Banks ─────────────────────────────────────────────────────────")
    BANKS = [
        ("BSA", "BANCO SANTANDER, S.A.",       "ES"),
        ("BST", "BANCO SANTANDER TOTTA",       "PT"),
        ("SUK", "SANTANDER UK",                "UK"),
        ("SCB", "SANTANDER CONSUMER BANK AG",  "DE"),
    ]
    exist_codes = _existing_set("banks", "code")
    _upsert("banks", ["code", "name", "country", "is_active"],
            [(c, n, co, 1) for c, n, co in BANKS if c not in exist_codes], "banks")

    print("── Products ──────────────────────────────────────────────────────")
    PRODUCTS = [
        ("CRED_IMP", "Créditos Importación"),
        ("REM_IMP",  "Remesas Importación"),
        ("REM_EXP",  "Remesas Exportación"),
        ("GAR_EMI",  "Garantias Emitidas"),
        ("GAR_REC",  "Garantias Recibidas"),
        ("ORD_PAG",  "Ordenes de Pago Financiadas"),
        ("EURO",     "Eurocobros"),
    ]
    exist_prod = _existing_set("products", "code")
    _upsert("products", ["code", "name", "is_active"],
            [(c, n, 1) for c, n in PRODUCTS if c not in exist_prod], "products")

    print("── Departments ───────────────────────────────────────────────────")
    DEPARTMENTS = [
        "OPAGO", "CDI SAN", "REM EXPORT", "FINANCIACIONES", "REM IMPORT",
        "CONTROL", "GARANTIAS", "GENERAL", "CDE", "PRUEBAS",
        "CHEQUES EXPORT", "PAGOS", "ESPANA", "ONE-UK TRADE",
        "PROYECTOS", "CONFIRMING", "FACTURACION",
    ]
    _insert_names("departments", DEPARTMENTS, label="departments")

    print("── Teams ─────────────────────────────────────────────────────────")
    TEAMS = [
        "CLIENT MANAGER", "RESPONSABLE", "DATA", "PROYECTOS",
        "METODOLOGIA Y SOPORTE", "PAGOS FINANCIADOS", "REMESAS",
        "CREDITOS DOCUMENTARIOS", "GARANTIAS", "Sourcing",
    ]
    _insert_names("teams", TEAMS, label="teams")

    print("── Error Impacts ─────────────────────────────────────────────────")
    exist_imp = _existing_set("error_impacts")
    _upsert("error_impacts", ["name", "level", "is_active"],
            [(n, lv, 1) for n, lv in [("Alto", "HIGH"), ("Bajo", "LOW")] if n not in exist_imp],
            "error_impacts")

    print("── Error Detected By ─────────────────────────────────────────────")
    _insert_names("error_detected_by", [
        "Corresponsal", "Cliente Final", "TRADE",
        "Of/Uni/Middle/Gestor", "Quality Unit",
    ], label="error_detected_by")

    print("── Error Origins ─────────────────────────────────────────────────")
    _insert_names("error_origins",
                  ["Trade_Personas", "Trade_Procesos", "Trade_Tecnologia", "Terceros"],
                  label="error_origins")

    print("── Tutoria Error Categories ──────────────────────────────────────")
    _insert_names("tutoria_error_categories", [
        "Formación", "Nuevo Conocimiento", "Interpretación y Análisis",
        "Omisión", "Marcación", "Organización", "Tipográfico",
        "Incidencia Cliente", "-- por clasificar --", "Metodologia",
        "Conocimientos", "Procedimientos", "Atención al detalle",
    ], label="tutoria_error_categories (tipología erro)")

    # Categories linked to origins
    cur.execute("SELECT name, id FROM error_origins")
    origin_map: dict[str, int] = {r[0]: r[1] for r in cur.fetchall()}
    exist_cat = _existing_set("tutoria_error_categories")
    TIPOLOGIA_INCIDENTE = [
        ("Formación Insuficiente",                    "Trade_Personas"),
        ("Dependencia de personal clave",             "Trade_Personas"),
        ("Error Puntual",                             "Trade_Personas"),
        ("Sobrecarga Operativa",                      "Trade_Personas"),
        ("Segregación Funcional",                     "Trade_Personas"),
        ("Diseño ineficaz del proceso",               "Trade_Procesos"),
        ("Desempeño ineficaz de un proceso",          "Trade_Procesos"),
        ("Calidad de los datos",                      "Trade_Procesos"),
        ("Gestión del cambio tecnológico inadecuado", "Trade_Tecnologia"),
        ("Diseño inadecuado de los sistemas",         "Trade_Tecnologia"),
        ("Funcionamiento inadecuado de un sistema",  "Trade_Tecnologia"),
        ("Proveedores",                               "Terceros"),
        ("Oficina/Uni/Middle",                        "Terceros"),
        ("Corresponsal",                              "Terceros"),
    ]
    rows_inc = [
        (name, origin_map[orig], 1)
        for name, orig in TIPOLOGIA_INCIDENTE
        if name not in exist_cat and orig in origin_map
    ]
    _upsert("tutoria_error_categories", ["name", "origin_id", "is_active"],
            rows_inc, "tutoria_error_categories (tipología incidente)")

    # ── SECÇÃO 2: Excel ────────────────────────────────────────────────────

    if not args.skip_excel:
        try:
            import openpyxl
        except ImportError:
            sys.exit("ERROR: pip install openpyxl")

        if not XLSX_FILE.exists():
            print(f"  [WARN] Excel não encontrado: {XLSX_FILE} — a saltar activities/error_types")
        else:
            print("── Activities (Excel TBL_ACTIVIDAD) ──────────────────────────────")
            wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)
            names_act = sorted({
                str(r[0]).strip()
                for r in wb["TBL_ACTIVIDAD"].iter_rows(min_row=2, values_only=True)
                if r[0] and str(r[0]).strip()
            })
            print(f"  {len(names_act)} actividades únicas no Excel")
            _insert_names("activities", names_act, label="activities")

            print("── Error Types (Excel TBL_TIPOERROR) ─────────────────────────────")
            names_et = sorted({
                str(r[4]).strip()
                for r in wb["TBL_TIPOERROR"].iter_rows(min_row=2, values_only=True)
                if r[4] and str(r[4]).strip()
            })
            wb.close()
            print(f"  {len(names_et)} tipos de erro únicos no Excel")
            _insert_names("error_types", names_et, label="error_types")

    # ── SECÇÃO 3: Utilizadores de teste ────────────────────────────────────

    print("── Test Users (password: test123) ────────────────────────────────")
    # bcrypt hash of "test123"
    HASH_TEST123 = "$2b$12$VbQ4uFFIK/Q.RInCdJVQ7e9rZDoTZOivTmXavig9mVxtyGRjLbZsK"

    cur.execute("SELECT id FROM teams LIMIT 1")
    team_row = cur.fetchone()
    team_id_db = team_row[0] if team_row else None

    # (email, full_name, role, is_trainer, is_tutor, is_liberador, is_team_lead, is_referente)
    # Roles disponíveis no portal: TRAINEE (base), MANAGER (chefe equipa), GESTOR, ADMIN
    # is_trainer=True num TRAINEE dá privilégios de formador (via require_role no backend)
    # Formato: (email, nome, role, is_trainer, is_tutor, is_liberador, is_team_lead, is_referente)
    CORE_USERS = [
        ("admin@tradehub.com",     "Admin Portal",     "ADMIN",    0, 0, 0, 0, 0),
        ("trainer@tradehub.com",   "Test Trainer",     "TRAINEE",  1, 0, 0, 0, 0),  # TRAINEE + is_trainer
        ("tutor@tradehub.com",     "Test Tutor",       "TRAINEE",  1, 1, 0, 0, 0),  # TRAINEE + is_trainer + is_tutor
        ("manager@tradehub.com",   "Test Manager",     "MANAGER",  0, 0, 0, 1, 0),
        ("student@tradehub.com",   "Test Student",     "TRAINEE",  0, 0, 0, 0, 0),
        ("liberador@tradehub.com", "Test Liberador",   "TRAINEE",  0, 0, 1, 0, 0),
        ("referente@tradehub.com", "Test Referente",   "TRAINEE",  0, 0, 0, 0, 1),
        ("chefe@tradehub.com",     "Test Chefe",       "MANAGER",  0, 0, 0, 1, 0),
        # CI / test-suite users
        ("manager_test@tradehub.com",   "Manager Test",   "MANAGER",  0, 0, 0, 1, 0),
        ("trainer_test@tradehub.com",   "Trainer Test",   "TRAINEE",  1, 0, 0, 0, 0),  # TRAINEE + is_trainer
        ("tutor_test@tradehub.com",     "Tutor Test",     "TRAINEE",  0, 1, 0, 0, 0),
        ("student_test@tradehub.com",   "Student Test",   "TRAINEE",  0, 0, 0, 0, 0),
        ("liberador_test@tradehub.com", "Liberador Test", "TRAINEE",  0, 0, 1, 0, 0),
        ("referente_test@tradehub.com", "Referente Test", "TRAINEE",  0, 0, 0, 0, 1),
        ("chefe_test@tradehub.com",     "Chefe Test",     "MANAGER",  0, 0, 0, 1, 0),
    ]

    exist_emails = _existing_set("users", "email")
    user_rows = [
        (email, name, HASH_TEST123, role, 1, 0, is_tr, is_tu, is_li, is_tl, is_re, team_id_db)
        for email, name, role, is_tr, is_tu, is_li, is_tl, is_re in CORE_USERS
        if email not in exist_emails
    ]
    _upsert("users",
            ["email", "full_name", "hashed_password", "role", "is_active", "is_pending",
             "is_trainer", "is_tutor", "is_liberador", "is_team_lead", "is_referente", "team_id"],
            user_rows, "users")

    # Assign tutor to student, liberador, referente
    if not DRY:
        cur.execute("SELECT id FROM users WHERE email = 'tutor@tradehub.com'")
        tr = cur.fetchone()
        if tr:
            cur.execute(
                "UPDATE users SET tutor_id = %s "
                "WHERE email IN ('student@tradehub.com','liberador@tradehub.com','referente@tradehub.com') "
                "AND tutor_id IS NULL",
                (tr[0],)
            )
            conn.commit()
            if cur.rowcount:
                print(f"  [OK] tutor_id={tr[0]} atribuído a {cur.rowcount} utilizador(es)")

        # Fix NULL is_active on banks/products (migration quirk)
        cur.execute("UPDATE banks SET is_active = 1 WHERE is_active IS NULL")
        cur.execute("UPDATE products SET is_active = 1 WHERE is_active IS NULL")
        conn.commit()

    # ── Summary fase 1 ─────────────────────────────────────────────────────
    print(f"\n{'─'*65}")
    SUMMARY_TABLES = [
        "users", "banks", "products", "departments", "teams",
        "error_impacts", "error_detected_by", "error_origins",
        "tutoria_error_categories", "activities", "error_types",
    ]
    total_m = 0
    for t in SUMMARY_TABLES:
        cur.execute(f"SELECT COUNT(*) FROM `{t}`")
        n = cur.fetchone()[0]
        print(f"  {t:<35} {n:>5} linhas")
        total_m += n
    print(f"  {'TOTAL DADOS-MESTRE':<35} {total_m:>5} linhas")
    print(f"{'─'*65}\n")

    cur.close()
    conn.close()

    if args.skip_operational:
        print("  [skip] Fase 2 (dados operacionais) saltada.\n")
        sys.exit(0)

else:
    print("  [skip] Fase 1 (dados-mestre) saltada.\n")


# ══════════════════════════════════════════════════════════════════════════════
# FASE 2 — DADOS OPERACIONAIS (via API)
# ══════════════════════════════════════════════════════════════════════════════

try:
    import requests
except ImportError:
    sys.exit("ERROR: pip install requests")

BASE = args.url.rstrip("/")
PASS_COUNT = 0
FAIL_COUNT = 0
CREATED: dict[str, list] = {}

TEST_PASSWORD = "test123"
USERS_API = {
    "ADMIN":     ("admin@tradehub.com",     TEST_PASSWORD),
    "TRAINER":   ("trainer@tradehub.com",   TEST_PASSWORD),
    "STUDENT":   ("student@tradehub.com",   TEST_PASSWORD),
    "MANAGER":   ("manager@tradehub.com",   TEST_PASSWORD),
    "TUTOR":     ("tutor@tradehub.com",     TEST_PASSWORD),
    "LIBERADOR": ("liberador@tradehub.com", TEST_PASSWORD),
}

TOKENS:   dict[str, str]  = {}
HEADERS:  dict[str, dict] = {}
USER_IDS: dict[str, int]  = {}


# ── API helpers ────────────────────────────────────────────────────────────

def _login(role: str, email: str, password: str) -> bool:
    for attempt in range(4):
        try:
            resp = requests.post(
                f"{BASE}/api/auth/login",
                data={"username": email, "password": password},
                timeout=15,
            )
        except requests.exceptions.ConnectionError:
            print(f"\n  FATAL: não é possível ligar a {BASE}")
            sys.exit(1)
        if resp.status_code == 200:
            TOKENS[role] = resp.json().get("access_token", "")
            HEADERS[role] = {"Authorization": f"Bearer {TOKENS[role]}"}
            return True
        elif resp.status_code == 429:
            wait = 15 * (attempt + 1)
            print(f"  [429] Rate limit para {role} — aguardar {wait}s")
            time.sleep(wait)
        else:
            print(f"  [WARN] Login falhou para {role}: {resp.status_code}")
            return False
    return False


def _get_me(role: str) -> int | None:
    resp = requests.get(f"{BASE}/api/auth/me", headers=HEADERS[role], timeout=10)
    if resp.status_code == 200:
        uid = resp.json().get("id")
        USER_IDS[role] = uid
        return uid
    return None


def _call(method: str, role: str, label: str, path: str,
          payload: dict | None = None, expected: int = 200) -> dict | None:
    global PASS_COUNT, FAIL_COUNT
    url = BASE + path
    try:
        if method == "POST":
            resp = requests.post(url, headers=HEADERS[role], json=payload, timeout=15)
        elif method == "PATCH":
            resp = requests.patch(url, headers=HEADERS[role], json=payload, timeout=15)
        elif method == "PUT":
            resp = requests.put(url, headers=HEADERS[role], json=payload, timeout=15)
        else:
            resp = requests.get(url, headers=HEADERS[role], timeout=10)
    except Exception as e:
        FAIL_COUNT += 1
        print(f"    X FAIL [{label}] {method} {path} — {e}")
        return None

    code = resp.status_code
    ok = code == expected or (code in (200, 201) and expected in (200, 201))
    try:
        body = resp.json()
    except Exception:
        body = {}

    if ok:
        PASS_COUNT += 1
        if V:
            eid = body.get("id", "") if isinstance(body, dict) else ""
            print(f"    + [{code}] {method} {path:<50} {f'id={eid}' if eid else ''}")
    else:
        FAIL_COUNT += 1
        detail = body.get("detail", str(body)[:120]) if isinstance(body, dict) else ""
        print(f"    X FAIL [{code}] {method} {path:<50} expected={expected} — {detail}")
        return None
    return body


def _post(role, label, path, payload, expected=200):
    return _call("POST", role, label, path, payload, expected)

def _patch(role, label, path, payload, expected=200):
    return _call("PATCH", role, label, path, payload, expected)

def _put(role, label, path, payload, expected=200):
    return _call("PUT", role, label, path, payload, expected)

def _get(role: str, path: str) -> dict | list | None:
    try:
        resp = requests.get(BASE + path, headers=HEADERS[role], timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None

def section(title: str) -> None:
    print(f"\n  -- {title} {'─' * max(1, 55 - len(title))}")

def today_str(offset_days: int = 0) -> str:
    return (datetime.now() + timedelta(days=offset_days)).strftime("%Y-%m-%d")


# ── 0. Login ───────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print(f"  PortalTradeHub — Seed Dados Operacionais via API")
print(f"  Target: {BASE}")
print(f"{'='*65}")

section("Login")
admin_ok = _login("ADMIN", *USERS_API["ADMIN"])
if not admin_ok:
    print("  FATAL: Admin login falhou. A terminar.")
    sys.exit(1)
print("  [OK] ADMIN")

for role in ["TRAINER", "STUDENT", "MANAGER", "TUTOR"]:
    time.sleep(2)
    ok = _login(role, *USERS_API[role])
    print(f"  [{'OK' if ok else 'FAIL'}] {role}")

elapsed = 2 * 4
wait_rl = max(1, 62 - elapsed)
print(f"  [..] Aguardar {wait_rl}s (rate limit)...")
time.sleep(wait_rl)

ok = _login("LIBERADOR", *USERS_API["LIBERADOR"])
print(f"  [{'OK' if ok else 'FAIL'}] LIBERADOR")

for role in USERS_API:
    if role in HEADERS:
        _get_me(role)
        if V:
            print(f"  {role} → user_id={USER_IDS.get(role)}")


# ── 1. Lookup IDs ──────────────────────────────────────────────────────────
section("Lookup de IDs para FK")

banks     = _get("ADMIN", "/api/admin/banks") or []
bank_ids  = [b["id"] for b in banks[:2]]
bank_id   = bank_ids[0] if bank_ids else None
print(f"  Banks: {len(banks)}, usando IDs={bank_ids}")

products     = _get("ADMIN", "/api/admin/products") or []
product_ids  = [p["id"] for p in products[:3]]
product_id   = product_ids[0] if product_ids else None
print(f"  Products: {len(products)}, usando IDs={product_ids}")

teams_api = _get("ADMIN", "/api/teams") or []
team_id   = teams_api[0]["id"] if teams_api else None
print(f"  Teams: {len(teams_api)}, team_id={team_id}")

impacts      = _get("ADMIN", "/api/admin/master/impacts") or []
impact_id    = impacts[0]["id"] if impacts else None
origins      = _get("ADMIN", "/api/admin/master/origins") or []
origin_id    = origins[0]["id"] if origins else None
detected_by  = _get("ADMIN", "/api/admin/master/detected-by") or []
detected_by_id = detected_by[0]["id"] if detected_by else None
departments  = _get("ADMIN", "/api/admin/master/departments") or []
department_id = departments[0]["id"] if departments else None

_raw_act = _get("ADMIN", "/api/admin/master/activities") or []
activities  = (_raw_act if isinstance(_raw_act, list)
               else _raw_act.get("items", _raw_act.get("data", [])))
activity_id = activities[0]["id"] if activities else None

_raw_et = _get("ADMIN", "/api/admin/master/error-types") or []
error_types = (_raw_et if isinstance(_raw_et, list)
               else _raw_et.get("items", _raw_et.get("data", [])))
error_type_id = error_types[0]["id"] if error_types else None

_raw_cat = _get("ADMIN", "/api/tutoria/categories") or []
categories  = (_raw_cat if isinstance(_raw_cat, list)
               else _raw_cat.get("items", _raw_cat.get("data", [])))
category_id = categories[0]["id"] if isinstance(categories, list) and categories else None

print(f"  impacts={len(impacts)}, origins={len(origins)}, detectedBy={len(detected_by)}")
print(f"  departments={len(departments)}, activities={len(activities)}, errorTypes={len(error_types)}")
print(f"  categories={len(categories) if isinstance(categories, list) else '?'}")


# ── 2. Admin: Team Services + FAQs ────────────────────────────────────────
section("ADMIN: Team Services")
if team_id and product_ids:
    for pid in product_ids[:2]:
        _post("ADMIN", "Team service", f"/api/teams/{team_id}/services", {"product_id": pid})

section("ADMIN: Chat FAQs")
FAQS = [
    {
        "keywords_pt": "plano formacao inscricao",
        "keywords_es": "plan formacion inscripcion",
        "keywords_en": "training plan enrollment",
        "answer_pt": "Para se inscrever num plano de formacao, aceda a 'Planos de Formacao' e clique em 'Inscrever-se'.",
        "answer_es": "Para inscribirse en un plan de formacion, acceda a 'Planes de Formacion' y haga clic en 'Inscribirse'.",
        "answer_en": "To enroll in a training plan, go to 'Training Plans' and click 'Enroll'.",
        "priority": 10,
    },
    {
        "keywords_pt": "erro tutoria reportar",
        "keywords_es": "error tutoria reportar",
        "keywords_en": "tutoring error report",
        "answer_pt": "Para reportar um erro de tutoria, va a 'Tutoria > Erros' e clique 'Novo Erro'.",
        "answer_es": "Para reportar un error de tutoria, vaya a 'Tutoria > Errores' y haga clic en 'Nuevo Error'.",
        "answer_en": "To report a tutoring error, go to 'Tutoring > Errors' and click 'New Error'.",
        "priority": 8,
    },
    {
        "keywords_pt": "certificado download descarregar",
        "keywords_es": "certificado descarga descargar",
        "keywords_en": "certificate download",
        "answer_pt": "Os certificados ficam disponiveis apos a conclusao de um plano. Aceda a 'Os Meus Certificados'.",
        "answer_es": "Los certificados estan disponibles tras completar un plan. Acceda a 'Mis Certificados'.",
        "answer_en": "Certificates are available after completing a training plan. Go to 'My Certificates'.",
        "priority": 7,
    },
    {
        "keywords_pt": "chamado ticket suporte",
        "keywords_es": "ticket soporte ayuda",
        "keywords_en": "ticket support help",
        "answer_pt": "Para abrir um chamado, aceda a 'Chamados' e clique 'Novo Chamado'.",
        "answer_es": "Para abrir un ticket, acceda a 'Tickets' y haga clic en 'Nuevo Ticket'.",
        "answer_en": "To open a support ticket, go to 'Tickets' and click 'New Ticket'.",
        "priority": 6,
    },
]
for faq in FAQS:
    _post("ADMIN", "FAQ", "/api/chat/faqs", faq)


# ── 3. Trainer: Cursos + Lições ────────────────────────────────────────────
section("TRAINER: Cursos")

COURSES = [
    {
        "title": "Creditos Documentarios - Nivel Basico",
        "description": "Formacao basica sobre creditos documentarios de importacao e exportacao.",
        "level": "BEGINNER", "course_type": "CURSO",
        "bank_ids": bank_ids[:1], "product_ids": product_ids[:1],
    },
    {
        "title": "Remesas Internacionales - Nivel Intermedio",
        "description": "Formacao sobre remessas de importacao e exportacao. Cobrancas documentarias.",
        "level": "INTERMEDIATE", "course_type": "CURSO",
        "bank_ids": bank_ids, "product_ids": product_ids[1:3] if len(product_ids) > 1 else product_ids,
    },
    {
        "title": "Garantias Bancarias - Nivel Avanzado",
        "description": "Formacao avancada sobre garantias emitidas e recebidas. URDG 758.",
        "level": "EXPERT", "course_type": "CURSO",
        "bank_ids": bank_ids, "product_ids": product_ids,
    },
]
course_ids = []
for c in COURSES:
    body = _post("TRAINER", f"Curso: {c['title'][:30]}", "/api/trainer/courses", c)
    if body and isinstance(body, dict) and body.get("id"):
        course_ids.append(body["id"])
CREATED["courses"] = course_ids

section("TRAINER: Lições")
LESSON_TEMPLATES = [
    [
        ("Introduccion a Creditos Documentarios", "Conceptos basicos, partes intervinientes y marco legal.", "THEORETICAL", 45),
        ("Tipos de Creditos Documentarios",       "Credito irrevocable, confirmado, transferible, back-to-back.", "THEORETICAL", 60),
        ("Operativa de Apertura de Credito",      "Proceso paso a paso para la apertura de un credito.", "PRACTICAL", 90),
        ("Documentos Requeridos",                 "Factura, BL, certificado de origen, packing list, seguro.", "THEORETICAL", 45),
    ],
    [
        ("Remesas de Importacion",  "Operativa completa de remesas simples y documentarias.", "THEORETICAL", 50),
        ("Remesas de Exportacion",  "Operativa completa de remesas de exportacion.", "THEORETICAL", 50),
        ("Practica: Alta de Remesa", "Ejercicio practico de alta de remesa en sistema.", "PRACTICAL", 120),
    ],
    [
        ("Marco Legal de Garantias", "URDG 758, tipos de garantias y normativa aplicable.", "THEORETICAL", 60),
        ("Garantias Emitidas",       "Proceso de emision, analisis de riesgo y documentacion.", "THEORETICAL", 45),
        ("Garantias Recibidas",      "Verificacion, reclamaciones y ejecucion.", "THEORETICAL", 45),
        ("Caso Practico: Garantia de Licitacion", "Resolver un caso completo de garantia.", "PRACTICAL", 90),
    ],
]
lesson_ids = []
for idx, cid in enumerate(course_ids):
    for order, (title, desc, ltype, mins) in enumerate(LESSON_TEMPLATES[idx % len(LESSON_TEMPLATES)]):
        body = _post("TRAINER", f"Lição: {title[:25]}", "/api/trainer/lessons", {
            "title": title, "description": desc, "lesson_type": ltype,
            "started_by": "TRAINER", "order_index": order,
            "estimated_minutes": mins, "course_id": cid,
        })
        if body and isinstance(body, dict) and body.get("id"):
            lesson_ids.append(body["id"])
CREATED["lessons"] = lesson_ids


# ── 4. Trainer: Planos de Formação ────────────────────────────────────────
section("TRAINER: Planos de Formação")

student_id = USER_IDS.get("STUDENT")
trainer_id = USER_IDS.get("TRAINER")

PLANS = [
    {
        "title": "Plan Formacion Q1 2026 - Creditos Documentarios",
        "description": "Plan trimestral para el equipo de creditos documentarios.",
        "trainer_ids": [trainer_id] if trainer_id else [],
        "student_ids": [student_id] if student_id else [],
        "bank_ids": bank_ids[:1], "product_ids": product_ids[:1],
        "start_date": today_str(-30), "end_date": today_str(60),
        "course_ids": course_ids[:1],
    },
    {
        "title": "Plan Formacion Remesas - Onboarding Nuevos",
        "description": "Onboarding para nuevos miembros del equipo de remesas.",
        "trainer_ids": [trainer_id] if trainer_id else [],
        "student_ids": [student_id] if student_id else [],
        "bank_ids": bank_ids,
        "product_ids": product_ids[1:3] if len(product_ids) > 1 else product_ids,
        "start_date": today_str(-15), "end_date": today_str(45),
        "course_ids": course_ids[1:2],
    },
    {
        "title": "Plan Formacion Garantias - Certificacion Avanzada",
        "description": "Certificacion avanzada para especialistas en garantias.",
        "trainer_ids": [trainer_id] if trainer_id else [],
        "bank_ids": bank_ids, "product_ids": product_ids,
        "start_date": today_str(0), "end_date": today_str(90),
        "course_ids": course_ids[2:3],
    },
]
plan_ids = []
for p in PLANS:
    body = _post("TRAINER", f"Plano: {p['title'][:30]}", "/api/training-plans/", p)
    if body and isinstance(body, dict) and body.get("id"):
        plan_ids.append(body["id"])
CREATED["plans"] = plan_ids

if plan_ids and len(plan_ids) >= 3 and student_id:
    _post("TRAINER", "Inscrever aluno plano 3",
          f"/api/training-plans/{plan_ids[2]}/assign",
          {"student_id": student_id, "start_date": today_str(0), "end_date": today_str(90)})


# ── 5. Trainer: Desafios ───────────────────────────────────────────────────
section("TRAINER: Desafios")

challenge_ids = []
if course_ids:
    CHALLENGES = [
        {
            "course_id": course_ids[0],
            "title": "Desafio Creditos - Operaciones Basicas",
            "description": "Realizar 10 operaciones de apertura de credito.",
            "difficulty": "easy", "challenge_type": "SUMMARY",
            "operations_required": 10, "time_limit_minutes": 60,
            "target_mpu": 6.0, "max_errors": 2,
        },
        {
            "course_id": course_ids[1] if len(course_ids) > 1 else course_ids[0],
            "title": "Desafio Remesas - Velocidad y Precision",
            "description": "Procesar 15 remesas con un maximo de 1 error.",
            "difficulty": "medium", "challenge_type": "SUMMARY",
            "operations_required": 15, "time_limit_minutes": 90,
            "target_mpu": 6.0, "max_errors": 1,
        },
        {
            "course_id": course_ids[-1],
            "title": "Desafio Garantias - Certificacion",
            "description": "Completar 8 operaciones de garantias sin errores.",
            "difficulty": "hard", "challenge_type": "COMPLETE",
            "operations_required": 8, "time_limit_minutes": 120,
            "target_mpu": 15.0, "max_errors": 0,
        },
    ]
    for ch in CHALLENGES:
        body = _post("TRAINER", f"Desafio: {ch['title'][:25]}", "/api/challenges/", ch)
        if body and isinstance(body, dict) and body.get("id"):
            challenge_ids.append(body["id"])
    for cid in challenge_ids:
        _post("TRAINER", "Libertar desafio", f"/api/challenges/{cid}/release", {"is_released": True})
    CREATED["challenges"] = challenge_ids

if challenge_ids and student_id:
    _post("TRAINER", "Submeter resultado desafio", "/api/challenges/submit/summary", {
        "challenge_id": challenge_ids[0], "user_id": student_id,
        "submission_type": "SUMMARY", "total_operations": 10,
        "total_time_minutes": 52.5, "errors_count": 1,
        "error_methodology": 0, "error_knowledge": 1,
        "error_detail": 0, "error_procedure": 0,
    })


# ── 6. Tutor: Erros de Tutoria ─────────────────────────────────────────────
section("TUTOR: Erros de Tutoria")

tutoria_error_ids = []
if "TUTOR" in HEADERS:
    TUTORIA_ERRORS = [
        {
            "date_occurrence": today_str(-10),
            "description": "Error en el tipo de cambio en remesa de importacion. Se uso el tipo del dia anterior.",
            "tutorado_id": student_id, "bank_id": bank_id, "impact_id": impact_id,
            "origin_id": origin_id, "category_id": category_id,
            "detected_by_id": detected_by_id, "department_id": department_id,
            "activity_id": activity_id, "error_type_id": error_type_id,
            "product_id": product_id, "severity": "ALTA",
            "reference_code": "REM-2026-00142", "currency": "USD", "amount": 25000.00,
            "office": "Madrid Central",
            "solution": "Se corrigio el tipo de cambio y se reproceso la operacion.",
            "motivos": [{"typology": "KNOWLEDGE", "description": "Desconocimiento del proceso de consulta de tipo de cambio"}],
        },
        {
            "date_occurrence": today_str(-7),
            "description": "Omision del campo 'Finalidad' en apertura de credito documentario.",
            "tutorado_id": student_id, "bank_id": bank_id, "impact_id": impact_id,
            "category_id": category_id, "detected_by_id": detected_by_id,
            "department_id": department_id, "severity": "MEDIA",
            "reference_code": "CDI-2026-00088", "currency": "EUR", "amount": 150000.00,
            "office": "Barcelona Trade",
            "motivos": [
                {"typology": "DETAIL",     "description": "Falta de atencion al detalle en campo obligatorio"},
                {"typology": "PROCEDURE",  "description": "No se siguio el checklist de apertura"},
            ],
        },
        {
            "date_occurrence": today_str(-3),
            "description": "Error en la clasificacion del tipo de garantia (emitida vs recibida).",
            "tutorado_id": student_id,
            "bank_id": bank_ids[1] if len(bank_ids) > 1 else bank_id,
            "impact_id": impact_id, "severity": "ALTA",
            "reference_code": "GAR-2026-00215", "currency": "EUR", "amount": 500000.00,
            "office": "Lisboa Trade",
            "motivos": [{"typology": "KNOWLEDGE", "description": "Confusion entre tipos de garantia"}],
        },
        {
            "date_occurrence": today_str(-1),
            "description": "Retraso en envio SWIFT MT700 por no verificar horario de corte del corresponsal.",
            "tutorado_id": student_id, "bank_id": bank_id, "severity": "MEDIA",
            "reference_code": "SWF-2026-00331", "office": "Madrid Central",
            "motivos": [{"typology": "METHODOLOGY", "description": "No se consulto el horario de corte"}],
        },
    ]
    for err in TUTORIA_ERRORS:
        body = _post("TUTOR", "Tutoria error", "/api/tutoria/errors", err)
        if body and isinstance(body, dict) and body.get("id"):
            tutoria_error_ids.append(body["id"])
    CREATED["tutoria_errors"] = tutoria_error_ids


# Action Plans + items
section("TUTOR: Planos de Ação")

action_plan_ids = []
if tutoria_error_ids and student_id:
    for eid, pdata in zip(tutoria_error_ids[:2], [
        {
            "tutorado_id": student_id, "plan_type": "CORRECTIVO",
            "description": "Formacion en consulta de tipos de cambio.",
            "what": "Formacion especifica sobre tipos de cambio y fuentes de consulta",
            "why": "El tutorado desconocia el proceso de consulta del tipo de cambio",
            "how": "Sesion practica revisando el manual de tipos de cambio",
            "who": "Tutor + Tutorado", "when_deadline": today_str(14),
            "expected_result": "El tutorado consulta correctamente el tipo de cambio",
        },
        {
            "tutorado_id": student_id, "plan_type": "PREVENTIVO",
            "description": "Implementacion de checklist para apertura de creditos.",
            "what": "Revision del checklist de apertura de creditos",
            "why": "Se omitio un campo obligatorio por no seguir el checklist",
            "how": "Crear checklist digital y revisar cada paso",
            "who": "Tutor", "when_deadline": today_str(7),
            "side_by_side": True,
            "observation_date": today_str(5),
            "observation_notes": "Observacion side-by-side para verificar uso del checklist",
        },
    ]):
        body = _post("TUTOR", f"Plano de ação erro {eid}", f"/api/tutoria/errors/{eid}/plans", pdata)
        if body and isinstance(body, dict) and body.get("id"):
            action_plan_ids.append(body["id"])

    if action_plan_ids:
        _post("TUTOR", "Action item 1", f"/api/tutoria/plans/{action_plan_ids[0]}/items",
              {"type": "CORRETIVA", "description": "Revisar manual de tipos de cambio", "due_date": today_str(7)})
        _post("TUTOR", "Action item 2", f"/api/tutoria/plans/{action_plan_ids[0]}/items",
              {"type": "CORRETIVA", "description": "Realizar 5 operaciones supervisadas", "due_date": today_str(14)})
        if len(action_plan_ids) > 1:
            _post("TUTOR", "Action item 3", f"/api/tutoria/plans/{action_plan_ids[1]}/items",
                  {"type": "PREVENTIVA", "description": "Crear checklist digital de apertura", "due_date": today_str(5)})

    CREATED["action_plans"] = action_plan_ids

# Comments
section("TUTOR: Comentários")
if tutoria_error_ids:
    _post("TUTOR", "Comentário erro", f"/api/tutoria/errors/{tutoria_error_ids[0]}/comments",
          {"content": "Se ha revisado el caso. Necesita refuerzo en consulta de tipos de cambio."})
if tutoria_error_ids and "STUDENT" in HEADERS:
    _post("STUDENT", "Comentário aluno", f"/api/tutoria/errors/{tutoria_error_ids[0]}/comments",
          {"content": "Entendido, me comprometo a seguir el procedimiento de verificacion."})
if action_plan_ids:
    _post("TUTOR", "Comentário plano", f"/api/tutoria/plans/{action_plan_ids[0]}/comments",
          {"content": "Inicio del plan correctivo programado para esta semana."})

# Learning Sheets
section("TUTOR: Fichas de Aprendizagem")
learning_sheet_ids = []
if tutoria_error_ids and student_id:
    for eid, sh in zip(tutoria_error_ids[:2], [
        {
            "error_id": tutoria_error_ids[0],
            "title": "Consulta de Tipos de Cambio - Procedimiento Correcto",
            "error_summary": "Se aplico el tipo de cambio del dia anterior en una remesa.",
            "root_cause": "Desconocimiento del proceso de consulta del tipo de cambio vigente.",
            "correct_procedure": "1. Acceder a la pantalla de tipos de cambio. 2. Seleccionar divisa y fecha. 3. Verificar.",
            "key_learnings": "Siempre consultar el tipo de cambio ANTES de procesar.",
            "reference_material": "Manual de Tipos de Cambio v3.2, Seccion 4.1",
            "is_mandatory": True,
        },
        {
            "error_id": tutoria_error_ids[1],
            "title": "Checklist Apertura Creditos Documentarios",
            "error_summary": "Se omitio el campo 'Finalidad' obligatorio en la apertura.",
            "root_cause": "No se siguio el checklist de apertura establecido.",
            "correct_procedure": "Seguir paso a paso el checklist. Verificar TODOS los campos obligatorios.",
            "key_learnings": "El checklist existe por una razon. Todos los campos son obligatorios.",
        },
    ]):
        body = _post("TUTOR", f"Ficha: {sh['title'][:25]}", "/api/tutoria/learning-sheets", sh)
        if body and isinstance(body, dict) and body.get("id"):
            learning_sheet_ids.append(body["id"])
    CREATED["learning_sheets"] = learning_sheet_ids


# ── 7. Tutor: Cápsulas Formativas ─────────────────────────────────────────
section("TUTOR: Cápsulas Formativas")

capsula_ids = []
if "TUTOR" in HEADERS:
    CAPSULAS = [
        {
            "title": "Capsula: Verificacion de Documentos en Creditos",
            "description": "Metodologia rapida para verificar documentos en creditos documentarios.",
            "course_type": "CAPSULA_METODOLOGIA", "level": "INTERMEDIATE",
        },
        {
            "title": "Capsula: Horarios de Corte SWIFT",
            "description": "Referencia rapida de horarios de corte para mensajes SWIFT.",
            "course_type": "CAPSULA_FUNCIONALIDADE", "level": "BEGINNER",
        },
        {
            "title": "Capsula: URDG 758 Puntos Clave",
            "description": "Resumen de puntos clave de la normativa URDG 758.",
            "course_type": "CAPSULA_METODOLOGIA", "level": "EXPERT",
        },
    ]
    for cap in CAPSULAS:
        body = _post("TUTOR", f"Cápsula: {cap['title'][:25]}", "/api/tutoria/capsulas", cap)
        if body and isinstance(body, dict) and body.get("id"):
            capsula_ids.append(body["id"])
    CREATED["capsulas"] = capsula_ids


# ── 8. Sensos + Erros Internos ─────────────────────────────────────────────
section("TUTOR: Sensos")

senso_ids = []
if "TUTOR" in HEADERS:
    for s in [
        {"name": "Senso Semanal 2026-W10", "description": "Controlo semanal semana 10.",
         "start_date": today_str(-14), "end_date": today_str(-7)},
        {"name": "Senso Semanal 2026-W11", "description": "Controlo semanal semana 11.",
         "start_date": today_str(-7),  "end_date": today_str(0)},
        {"name": "Senso Semanal 2026-W12", "description": "Controlo semanal semana atual.",
         "start_date": today_str(0),   "end_date": today_str(7)},
    ]:
        body = _post("TUTOR", f"Senso: {s['name']}", "/api/internal-errors/sensos", s)
        if body and isinstance(body, dict) and body.get("id"):
            senso_ids.append(body["id"])
    CREATED["sensos"] = senso_ids

section("LIBERADOR: Erros Internos")
internal_error_ids = []
if "LIBERADOR" in HEADERS and senso_ids and student_id:
    INT_ERRORS = [
        {
            "senso_id": senso_ids[0], "gravador_id": student_id,
            "description": "Operacao de remesa processada com dados incorretos do beneficiario.",
            "reference_code": "IE-2026-001", "date_occurrence": today_str(-12),
            "impact_id": impact_id, "category_id": category_id,
            "department_id": department_id, "activity_id": activity_id, "bank_id": bank_id,
            "classifications": [{"classification": "KNOWLEDGE", "description": "Formato incorreto de dados do beneficiario"}],
            "peso_liberador": 3,
        },
        {
            "senso_id": senso_ids[0], "gravador_id": student_id,
            "description": "IBAN do ordenante incorreto em pagamento financiado.",
            "reference_code": "IE-2026-002", "date_occurrence": today_str(-11),
            "impact_id": impact_id, "department_id": department_id, "bank_id": bank_id,
            "classifications": [{"classification": "DETAIL", "description": "Erro tipografico no IBAN"}],
            "peso_liberador": 2,
        },
        {
            "senso_id": senso_ids[1] if len(senso_ids) > 1 else senso_ids[0],
            "gravador_id": student_id,
            "description": "Garantia emitida com data de vencimento incorreta (2025 em vez de 2026).",
            "reference_code": "IE-2026-003", "date_occurrence": today_str(-5),
            "impact_id": impact_id, "bank_id": bank_ids[1] if len(bank_ids) > 1 else bank_id,
            "classifications": [
                {"classification": "DETAIL",    "description": "Erro na data de vencimento"},
                {"classification": "PROCEDURE", "description": "Nao se verificou a data antes de confirmar"},
            ],
            "peso_liberador": 4,
        },
        {
            "senso_id": senso_ids[1] if len(senso_ids) > 1 else senso_ids[0],
            "gravador_id": student_id,
            "description": "Duplicacao de comissao em credito documentario de importacao.",
            "reference_code": "IE-2026-004", "date_occurrence": today_str(-4),
            "impact_id": impact_id, "department_id": department_id, "bank_id": bank_id,
            "peso_liberador": 5,
        },
        {
            "senso_id": senso_ids[-1],
            "gravador_id": student_id,
            "description": "Remesa de exportacao enviada sem documentacao de alfandega.",
            "reference_code": "IE-2026-005", "date_occurrence": today_str(-2),
            "department_id": department_id, "bank_id": bank_id,
            "classifications": [{"classification": "METHODOLOGY", "description": "Nao se seguiu o procedimento de envio"}],
            "peso_liberador": 3,
        },
    ]
    for ie in INT_ERRORS:
        body = _post("LIBERADOR", f"Erro interno: {ie['reference_code']}",
                     "/api/internal-errors/errors", ie)
        if body and isinstance(body, dict) and body.get("id"):
            internal_error_ids.append(body["id"])
    CREATED["internal_errors"] = internal_error_ids

section("TUTOR: Planos + Fichas de Erros Internos")
if internal_error_ids and "TUTOR" in HEADERS:
    _post("TUTOR", "IE plano de ação 1",
          f"/api/internal-errors/errors/{internal_error_ids[0]}/action-plan", {
              "description": "Plano correctivo para erros em dados do beneficiario",
              "items": [
                  {"description": "Rever formato correto de dados do beneficiario", "type": "CORRETIVA"},
                  {"description": "Criar guia rapida de formatos por pais", "type": "PREVENTIVA"},
              ],
          })
    # Use internal error #3 for learning sheet to avoid conflict with
    # tutoria learning sheets which use error_ids 1 and 2 in the shared table
    ie_ls_id = internal_error_ids[3] if len(internal_error_ids) > 3 else internal_error_ids[-1]
    _post("TUTOR", "IE ficha de aprendizagem 1",
          f"/api/internal-errors/errors/{ie_ls_id}/learning-sheet", {
              "error_summary": "Dados incorretos do beneficiario em remesa de importacao.",
              "impact_description": "Atraso de 24h no processamento por devolucao do corresponsal.",
              "actions_taken": "Corrigiram-se os dados e reprocessou-se a operacao.",
              "lessons_learned": "Verificar SEMPRE o formato de dados do beneficiario.",
              "recommendations": "Consultar tabela de formatos por pais antes de processar.",
              "error_weight": 3,
          })
    if len(internal_error_ids) > 2:
        _post("TUTOR", "IE plano de ação 2",
              f"/api/internal-errors/errors/{internal_error_ids[2]}/action-plan", {
                  "description": "Plano correctivo para erro de data em garantia",
                  "items": [{"description": "Implementar dupla verificacao de datas", "type": "CORRETIVA"}],
              })


# ── 9. Chamados de Suporte ─────────────────────────────────────────────────
section("Chamados (Support Tickets)")
chamado_ids = []
for role, ch in [
    ("STUDENT", {
        "title": "Error al acceder al modulo de certificados",
        "description": "Cuando intento acceder a 'Mis Certificados' aparece pantalla en blanco. Chrome y Firefox.",
        "type": "BUG", "priority": "ALTA", "portal": "FORMACOES",
    }),
    ("STUDENT", {
        "title": "Solicitud: Agregar filtro por fecha en historial de errores",
        "description": "Seria util poder filtrar errores de tutoria por rango de fechas.",
        "type": "MELHORIA", "priority": "MEDIA", "portal": "TUTORIA",
    }),
    ("MANAGER", {
        "title": "Dashboard de equipo no muestra datos del mes actual",
        "description": "Muestra estadisticas hasta el mes pasado pero no el mes actual.",
        "type": "BUG", "priority": "ALTA", "portal": "RELATORIOS",
    }),
    ("TRAINER", {
        "title": "Solicitud: Exportar lista de alumnos a Excel",
        "description": "Necesito exportar la lista de alumnos por plan en formato Excel.",
        "type": "MELHORIA", "priority": "MEDIA", "portal": "FORMACOES",
    }),
]:
    if role not in HEADERS:
        continue
    body = _post(role, f"Chamado: {ch['title'][:30]}", "/api/chamados", ch)
    if body and isinstance(body, dict) and body.get("id"):
        chamado_ids.append(body["id"])

if chamado_ids:
    _put("ADMIN", "Update chamado 1", f"/api/chamados/{chamado_ids[0]}",
         {"status": "EM_ANDAMENTO", "assigned_to_id": USER_IDS.get("ADMIN"),
          "admin_notes": "A investigar o problema. Parece ser cache do navegador."})
    _post("ADMIN", "Comentário chamado admin", f"/api/chamados/{chamado_ids[0]}/comments",
          {"content": "Identificado o problema. A trabalhar na solucao."})
    if "STUDENT" in HEADERS:
        _post("STUDENT", "Comentário chamado aluno", f"/api/chamados/{chamado_ids[0]}/comments",
              {"content": "Obrigado pela rapida resposta. Aguardo a solucao."})
CREATED["chamados"] = chamado_ids


# ── 10. Feedback Surveys + Respostas + Ações ──────────────────────────────
section("TUTOR: Feedback Surveys")
survey_ids = []
if "TUTOR" in HEADERS:
    for sv in [
        {"title": "Feedback Semanal - Semana 10",
         "week_start": today_str(-14), "week_end": today_str(-7)},
        {"title": "Feedback Semanal - Semana 11",
         "week_start": today_str(-7),  "week_end": today_str(0)},
    ]:
        body = _post("TUTOR", f"Survey: {sv['title'][:30]}", "/api/feedback/surveys", sv)
        if body and isinstance(body, dict) and body.get("id"):
            survey_ids.append(body["id"])
    CREATED["surveys"] = survey_ids

section("LIBERADOR: Feedback Responses")
response_ids = []
if survey_ids and "LIBERADOR" in HEADERS and student_id:
    for sv_id, resp_data in zip(survey_ids, [
        {
            "grabador_id": student_id, "opinion": "Melhoria esta semana. Menos erros em dados de beneficiario.",
            "sentiment": "POSITIVE",
            "concrete_situations": "REM-2026-00190 processada corretamente apos formacao.",
            "needs_tutor_intervention": False,
        },
        {
            "grabador_id": student_id, "opinion": "Erro pontual na data de garantia mas qualidade melhorou.",
            "sentiment": "NEUTRAL",
            "concrete_situations": "GAR-2026-00215 com erro de data mas detetado na revisao.",
            "needs_tutor_intervention": True,
        },
    ]):
        body = _post("LIBERADOR", "Feedback response", "/api/feedback/responses",
                     {"survey_id": sv_id, **resp_data})
        if body and isinstance(body, dict) and body.get("id"):
            response_ids.append(body["id"])
    CREATED["feedback_responses"] = response_ids

section("TUTOR: Feedback Actions")
if response_ids and "TUTOR" in HEADERS:
    _post("TUTOR", "Feedback action 1", "/api/feedback/actions", {
        "response_id": response_ids[0], "action_type": "FEEDBACK_DIRETO",
        "description": "Felicitar o operador pela melhoria em operacoes de remessas.",
    })
    if len(response_ids) > 1:
        _post("TUTOR", "Feedback action 2", "/api/feedback/actions", {
            "response_id": response_ids[1], "action_type": "TUTORIA",
            "description": "Sessao de reforco sobre verificacao de datas em garantias.",
        })


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY FINAL
# ══════════════════════════════════════════════════════════════════════════════
total_calls = PASS_COUNT + FAIL_COUNT
pct = int(PASS_COUNT / total_calls * 100) if total_calls else 0

print(f"\n{'='*65}")
print(f"  RESULTADOS: {PASS_COUNT}/{total_calls} chamadas API ({pct}%)")
print(f"{'='*65}")

print(f"\n  Dados operacionais criados:")
for entity, ids in CREATED.items():
    print(f"    {entity:<25} {len(ids):>3} registos  {ids}")

SUMMARY = {
    "Cursos":               len(CREATED.get("courses", [])),
    "Lições":               len(CREATED.get("lessons", [])),
    "Planos de Formação":   len(CREATED.get("plans", [])),
    "Desafios":             len(CREATED.get("challenges", [])),
    "Erros de Tutoria":     len(CREATED.get("tutoria_errors", [])),
    "Planos de Ação":       len(CREATED.get("action_plans", [])),
    "Fichas de Aprendizagem": len(CREATED.get("learning_sheets", [])),
    "Cápsulas Formativas":  len(CREATED.get("capsulas", [])),
    "Sensos":               len(CREATED.get("sensos", [])),
    "Erros Internos":       len(CREATED.get("internal_errors", [])),
    "Chamados":             len(CREATED.get("chamados", [])),
    "Feedback Surveys":     len(CREATED.get("surveys", [])),
    "Feedback Respostas":   len(CREATED.get("feedback_responses", [])),
    "FAQs":                 len(FAQS),
}
total_records = sum(SUMMARY.values())
print(f"\n  Resumo:")
for name, count in SUMMARY.items():
    print(f"    {name:<28} {count:>3}")
print(f"    {'TOTAL':<28} {total_records:>3}")

print(f"\n{'='*65}")
if FAIL_COUNT == 0:
    print(f"  Todas as {PASS_COUNT} chamadas passaram. Portal pronto.")
else:
    print(f"  {FAIL_COUNT} chamadas falharam. Ver erros acima.")
print(f"{'='*65}\n")

sys.exit(0 if FAIL_COUNT == 0 else 1)
